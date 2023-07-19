import dash
from dash import dcc
from dash import html
from dash.dependencies import Input, Output, ClientsideFunction, State
import dash_bootstrap_components as dbc
import numpy as np
import pandas as pd
import datetime
from datetime import datetime as dt
import pathlib
import plotly.graph_objects as go
from data_convert import create_shmoo_plot, read_shmoo_data
from test_shmoo_data import test_shmoo_data
import openpyxl
import io 
import base64
app = dash.Dash(
    __name__,
    external_stylesheets=["/assets/Bootstrap-css.css", "/assets/styles.css"]
    # meta_tags=[{"name": "viewport", "content": "width=device-width, initial-scale=1"}],
)
print(dbc.themes.BOOTSTRAP)
app.title = "Shmoo Analytics Dashboard"

server = app.server
app.config.suppress_callback_exceptions = True

# Path
BASE_PATH = pathlib.Path(__file__).parent.resolve()
DATA_PATH = BASE_PATH.joinpath("data").resolve()

def description_card():
    return dbc.Card(
        dbc.CardBody(
            [
                html.H5("Shmoo Reader"),
                html.H3("Welcome to the Shmoo Reader Dashboard"),
                html.P("Explore different Shmoo Files to Visualize and Generate Shmoo Plot"),
            ]
        )
    )


def generate_control_card():
    # ... Existing code ...
    product_options = [
        {"label": "Waikiki", "value": "Waikiki"},
        {"label": "Hamilton", "value": "Hamilton"},
        {"label": "Magnus", "value": "Magnus"},
    ]

    admit_options = [
        {"label": "TDF", "value": "TDF"},
        {"label": "ATPG", "value": "ATPG"},
        {"label": "MBIST", "value": "MBIST"},
    ]

    filetype_options = [
        {"label": "Multiple Files", "value": "Multiple Files"},
        {"label": "Multiple Folder", "value": "Multiple Folder"},
    ]
    return dbc.Card(
        dbc.CardBody(
            [
                dbc.Spinner(
                    [
                        html.Div(
                            [
                                html.Label("Select Product"),
                                dcc.Dropdown(id="product-select", options=product_options, value="Waikiki"),
                            ]
                        ),
                        html.Br(),
                        html.Div(
                            [
                                html.Label("Select Shmoo Test Type"),
                                dcc.Dropdown(id="admit-select", options=admit_options, value="TDF", multi=True),
                            ]
                        ),
                        html.Br(),
                        html.Div(
                            [
                                html.Label("Upload CSV"),
                                dcc.Upload(
                                    id="upload-data",
                                    children=html.Div(["Drag and drop or click to upload."]),
                                    style={
                                        "width": "100%",
                                        "height": "80px",
                                        "lineHeight": "60px",
                                        "borderWidth": "1px",
                                        "borderStyle": "dashed",
                                        "borderRadius": "5px",
                                        "textAlign": "center",
                                        "margin": "10px",
                                    },
                                    multiple=True,
                                ),
                            ]
                        ),
                        html.Br(),
                        dbc.Col(
                            [
                                dbc.Button("Analyze File", id="analyze-btn", n_clicks=0, color="primary", className="me-2"),
                                dbc.Button("Download", id="download-btn", n_clicks=0, color="primary", className="me-2"),
                                dbc.Button("Reset", id="reset-btn", n_clicks=0, color="danger", className="me-2"),
                                html.Label("", id="download_container")
                            ],
                            loading_state={
                                "is_loading": True,
                                "is_disabled": True
                            }
                        ),
                    ],
                    size="lg",
                    color="primary",
                    type="grow",
                )
            ]
        )
    )


def select_test_program_card():
    # ... Existing code ...

    return dbc.Card(
        dbc.CardBody(
            [
                dbc.Spinner(
                    [
                        html.Div(
                            [   
                                dbc.Col(
                                    [
                                        html.Label("Select Die: "),
                                        dcc.Dropdown(
                                            id="die_select",
                                            options=[{"label": i, "value": i} for i in ["Die"]],
                                            value="Die",
                                        ),

                                        html.Label("Select Plot Test Program:"),
                                        dcc.Dropdown(
                                            id="test_program_select",
                                            options=[{"label": i, "value": i} for i in ["TestProgram1"]],
                                            value="TestProgram1",
                                        ),
                                        html.Br(),
                                        html.Label("Test Info"),
                                        html.Div(id="test_info_div"),

                                    ]
                                )
                            ]
                        ),
                    ],
                    size="lg",
                    color="primary",
                    type="grow",
                )
            ]
        )
    )


def show_test_results():
    return html.Div(
    id="show_result_card"
        )



def show_heat_map():
    return  dbc.Card(
                    dbc.CardBody(
                        [
                            dbc.Spinner(
                                [
                                    html.Hr(),
                                    dcc.Graph(id="shmoo_plot_hm"),
                                ],                   
                                size="lg",
                                color="primary",
                                type="grow",
                            ),

                        ]
                        ),
                    )

app.layout = dbc.Container(
    [
        # Banner
        # dbc.Row(
        #     dbc.Col(html.Img(src=app.get_asset_url("logo.png")), className="banner"),
        # ),
        dcc.Store(id='Shmoo_Data_Dict'),
         
        dbc.Row(
            dbc.NavbarSimple(
                    children=[
                        dbc.NavItem(dbc.NavLink("Page 1", href="#")),
                        dbc.DropdownMenu(
                            children=[
                                dbc.DropdownMenuItem("More pages", header=True),
                            ],
                            nav=True,
                            in_navbar=True,
                            label="More",
                        ),
                    ],
                    brand="Navbar",
                    brand_href="#",
                    color="primary",
                    dark=True,
                ),
        ),

        html.Br(),
        
        dbc.Row(
            [
                # Left column
                dbc.Col(
                    [
                        description_card(),
                        html.Br(),
                        generate_control_card(),
                    ],
                    width=4,
                ),
                # Right column
                dbc.Col(
                    [
                        # Shmoo Plot Heatmap
                        show_heat_map(),


                        html.Br(),

                        # Plot Select Setting
                        dbc.Card(
                            dbc.CardBody(
                                [
                                    html.B("Shmoo Plot Setting"),
                                    html.Hr(),
                                    select_test_program_card(),
                                ]
                            )
                        ),
                        
                    ],
                    width=8,
                ),
            ]
        ),


   
    ],
    fluid=True,
)

@app.callback(
    Output("die_select", "options"),
    Output("die_select", "value"),
    Output("Shmoo_Data_Dict", "data"),
    # Output("test_program_select", "options"),
    [Input("analyze-btn", "n_clicks")]
)
def update_test_program_options(n_clicks):
    if n_clicks:
        
        result_dict = read_shmoo_data(test_shmoo_data)
        result_dict_keys = list(result_dict.keys())
        die_options = [{"label": key, "value": key} for key in result_dict_keys]
        die_value = result_dict_keys[0] if result_dict_keys else None

        if die_value:
            result_dict[die_value].keys()
            test_program_options = [{"label": key, "value": key} for key in result_dict[die_value].keys()]
        return die_options, die_value, result_dict, 
    return [], None, {}


@app.callback(
    Output("test_program_select", "options"),
    Output("test_program_select", "value"),
    [Input("die_select", "value"), 
     Input("Shmoo_Data_Dict", "data")],
     allow_duplicate=True
)
def update_test_program_options(die_value, result_dict):
    if die_value:
        sub_dict = result_dict.get(die_value)
        if sub_dict:
            result_value_keys = list( sub_dict.keys())
            test_program_options = [{"label": key, "value": key} for key in result_value_keys]
            test_program_value = result_value_keys[0] if result_value_keys else None
            text = result_dict[die_value][test_program_value]["test_info"]
            formatted_text = text.replace("\n", "  \n")
            return test_program_options, test_program_value
    return [], None


@app.callback(
    Output("shmoo_plot_hm", "figure"),
    Output("test_info_div", "children"),
    [Input("die_select", "value"),
     Input("test_program_select", "value"),
     Input("Shmoo_Data_Dict", "data")], 
     allow_duplicate=True
)
def update_selected_test_program(die_value, test_program_value, result_dict):
    if die_value:
        selected_test_data = result_dict[die_value][test_program_value]["shmoo_data"]
        fig = create_shmoo_plot(shmoo_data=selected_test_data)
        text = result_dict[die_value][test_program_value]["test_info"]
        formatted_text = text.replace("\n", "  \n")
        return fig, dcc.Markdown(children=formatted_text)

    return go.Figure(), ""


@app.callback(
    Output('download-btn', 'href'),
    [Input('download-btn', 'n_clicks'), 
    State("Shmoo_Data_Dict", 'data')] 
)
def generate_excel_file(n_clicks, result_dict):
    if n_clicks:
        # Create a BytesIO object to hold the Excel file contents
        excel_buffer = io.BytesIO()

        # Create the workbook
        workbook = openpyxl.Workbook()

        # Iterate over result_dict
        for die in result_dict:
            sheet_name = die 
            sheet = workbook.create_sheet(title=sheet_name)  # Create a new sheet with the desired name
            test_data = result_dict[die]
            current_row_index = 0  # Initialize the current row index

            # Iterate over test_program_name
            for test_program_name in test_data:
                test_info =  test_data[test_program_name]["test_info"]
                for cell_data in test_info.split("\n"):
                    # print(cell_data)
                    sheet.cell(row=current_row_index + 1, column=1).value = cell_data
                    current_row_index += 1
                shmoo_data = test_data[test_program_name]["shmoo_data"]
                print(shmoo_data)
                #Iterate over shmoo_data
                for index, row_data in enumerate(shmoo_data):
                    print(row_data)
                    for column2, cell_value in enumerate(row_data):
                        sheet.cell(row=index + current_row_index + 1, column=column2 + 1).value = cell_value
        
        # Save the workbook to the BytesIO object
        workbook.save(excel_buffer)

        # Set the BytesIO object position to the start of the stream
        excel_buffer.seek(0)

        # Generate a download link for the Excel file
        href = f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64.b64encode(excel_buffer.read()).decode()}'
        return href
    
    return ''
# @app.callback(
#     Output("patient_volume_hm", "figure"),
#     [Input("analyze-btn", "n_clicks")],
#     allow_duplicate=True
# )
# def generate_heatmap(n_clicks):
#     if n_clicks:
#         # Generate heatmap data
#         data = [[0.1, 0.3, 0.5, 0.7, 0.9],
#                 [1, 0.8, 0.6, 0.4, 0.2],
#                 [0.2, 0, 0.5, 0.7, 0.9],
#                 [0.9, 0.8, 0.4, 0.2, 0],
#                 [0.3, 0.4, 0.5, 0.7, 1]]

#         # fig = go.Figure(data=go.Heatmap(z=data))
#         # fig.update_layout(title="Shmoo Heatmap")
#         fig = create_shmoo_plot(shmoo_data=shmoo_data)

#         return fig

#     # Return an empty figure if the button has not been clicked
#     return go.Figure()


# @app.callback(
#     Output("patient_volume_hm", "figure"),
#     [Input("reset-btn", "n_clicks")],
#     allow_duplicate=True
# )
# def reset_plots(reset_clicks):
#     if reset_clicks:
#         # Return an empty figure to clear the plot
#         return go.Figure()

#     # Return the current figure if the button has not been clicked
#     return dash.no_update

# Run the server
if __name__ == "__main__":
    app.run_server(debug=True)